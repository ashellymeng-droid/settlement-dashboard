"""
歌单推广任务 — 结算引擎
支持平台: 小红书 / 抖音 / 快手 / B站 / 视频号
结算规则: 见 SETTLEMENT_RULES 常量
"""
import pandas as pd
import numpy as np
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

# ============================================================
# 结算规则配置
# ============================================================
SETTLEMENT_RULES = {
    "小红书": {
        "真人": {
            "cumulative": [(100, 150), (200, 300), (500, 600), (1000, 1000),
                          (2000, 1600), (3000, 2100), (5000, 3000)],
            "boom": [(5000, 3000), (10000, 4000)],
            "boom_threshold": 5000,
            "content_tag": "真人类歌单",
        },
        "图文": {
            "cumulative": [(100, 80), (200, 150), (500, 350), (1000, 500),
                          (2000, 1000), (3000, 1500), (5000, 2000)],
            "boom": [(5000, 2000), (10000, 3000)],
            "boom_threshold": 5000,
            "content_tag": "图文类歌单",
        },
        "视频": {
            "cumulative": [(200, 80), (500, 150), (1000, 300), (2000, 500),
                          (5000, 1000), (10000, 1200)],
            "boom": [(10000, 1200), (20000, 2000)],
            "boom_threshold": 10000,
            "content_tag": "视频类歌单",
        },
    },
    "抖音": {
        "真人": [(5000, 30), (10000, 60), (30000, 180), (50000, 300),
                (100000, 500), (200000, 1000), (300000, 1500), (500000, 2000)],
        "非真人": [(5000, 20), (10000, 40), (30000, 100), (50000, 150),
                 (100000, 300), (200000, 500), (300000, 700), (500000, 1000)],
        "max_posts": 10,
    },
    "快手": {
        "真人": [(5000, 25), (10000, 50), (30000, 135), (50000, 225),
                (100000, 400), (200000, 600), (300000, 900), (500000, 1200)],
        "非真人": [(5000, 15), (10000, 30), (30000, 75), (50000, 110),
                 (100000, 220), (200000, 420), (300000, 600), (500000, 800)],
        "max_posts": 10,
    },
    "B站/视频号": {
        "tiers": [(1000, 50), (10000, 200)],
    },
}


@dataclass
class SettlementConfig:
    """结算配置参数"""
    cap_per_person: int = 10000
    play_field: str = "7日播放量"
    play_fallback_field: str = "7日播放量(三方)"
    like_field: str = "7日点赞量"
    interact_field: str = "7日互动量"
    content_tag_field: str = "稿件内容标签"
    platform_field: str = "发布平台"
    review_field: str = "审核结果"
    access_field: str = "是否可访问(最新)"
    creator_id_field: str = "创作匠/易闪ID"
    creator_name_field: str = "创作匠/易闪昵称"
    post_id_field: str = "作品ID"
    netease_platform: str = "网易云音乐"


@dataclass
class CreatorDetail:
    """达人结算明细"""
    creator_id: str
    name: str
    total: float
    before_cap: float
    is_capped: bool
    platforms: Dict[str, float] = field(default_factory=dict)
    breakdown: List[dict] = field(default_factory=list)


@dataclass
class SettlementResult:
    """结算结果"""
    config: SettlementConfig
    creator_totals: Dict[str, Dict] = field(default_factory=dict)
    creator_details: Dict[str, CreatorDetail] = field(default_factory=dict)
    stats: Dict = field(default_factory=dict)
    raw_data: pd.DataFrame = None
    settled_posts: pd.DataFrame = None
    # 未分发网易云但本应结算的达人
    no_netease_totals: Dict[str, Dict] = field(default_factory=dict)
    no_netease_details: Dict[str, CreatorDetail] = field(default_factory=dict)
    no_netease_posts: pd.DataFrame = None


class SettlementEngine:
    """结算引擎"""

    def __init__(self, config: Optional[SettlementConfig] = None):
        self.config = config or SettlementConfig()
        self.result = SettlementResult(config=self.config)

    @staticmethod
    def apply_tier(value: float, tiers: List[Tuple[float, float]]) -> float:
        """阶梯定价：找到最高匹配档位"""
        amount = 0
        for threshold, reward in tiers:
            if value > threshold:
                amount = reward
        return amount

    @staticmethod
    def get_tier_label(value: float, tiers: List[Tuple[float, float]]) -> str:
        """获取命中的档位标签"""
        label = "未达标"
        for threshold, reward in tiers:
            if value > threshold:
                label = f">{threshold:,}→¥{reward}"
        return label

    def load_data(self, path: str) -> pd.DataFrame:
        """加载并清洗底表数据"""
        df = pd.read_excel(path, sheet_name='底表') if '底表' in \
            pd.ExcelFile(path).sheet_names else pd.read_excel(path)

        cfg = self.config
        df = df.copy()
        df.loc[:, cfg.access_field] = df[cfg.access_field].fillna('是')
        df.loc[:, '_can_settle'] = (
            (df[cfg.review_field] == '审核通过') &
            (df[cfg.access_field] == '是')
        )
        df.loc[:, cfg.content_tag_field] = df[cfg.content_tag_field].fillna('未分类')
        df.loc[:, '_play'] = df[cfg.play_field].fillna(df[cfg.play_fallback_field])

        # 标记有网易云分发的创作匠
        netease_creators = set(df.loc[
            df[cfg.platform_field] == cfg.netease_platform,
            cfg.creator_id_field
        ].unique())
        df.loc[:, '_has_netease'] = df[cfg.creator_id_field].isin(netease_creators)

        self.result.raw_data = df
        self.result.stats['total_rows'] = len(df)
        self.result.stats['netease_rows'] = int((df[cfg.platform_field] == cfg.netease_platform).sum())
        self.result.stats['netease_creators'] = len(netease_creators)

        return df

    def _settle_creators(self, posts_df, cfg, rules):
        """对一组稿件执行结算计算，返回 (creator_totals, creator_details)"""
        creator_totals = {}
        creator_details = {}

        for cid, group in posts_df.groupby(cfg.creator_id_field):
            name = group[cfg.creator_name_field].iloc[0]
            total = 0.0
            platforms = {}
            breakdown = []

            # 小红书
            xhs = group[group[cfg.platform_field] == '小红书']
            if len(xhs) > 0:
                xhs_total = 0.0
                for ct_name, ct_rules in rules['小红书'].items():
                    if ct_name in ['真人', '图文']:
                        tag = ct_rules['content_tag']
                        posts = xhs[xhs[cfg.content_tag_field] == tag]
                    else:
                        known_tags = [rules['小红书'][k]['content_tag'] for k in ['真人', '图文']]
                        posts = xhs[~xhs[cfg.content_tag_field].isin(known_tags)]
                    if len(posts) > 0:
                        sub = self._calc_xhs(posts, ct_rules['cumulative'], ct_rules['boom'],
                                            ct_rules['boom_threshold'], cfg, label=f"小红书-{ct_name}")
                        if sub['小计'] > 0:
                            xhs_total += sub['小计']
                            breakdown.append(sub)
                if xhs_total > 0:
                    platforms['小红书'] = xhs_total
                    total += xhs_total

            # 抖音
            dy = group[group[cfg.platform_field] == '抖音']
            if len(dy) > 0:
                dy_total = 0.0
                for ct_name in ['真人', '非真人']:
                    tag = rules['小红书']['真人']['content_tag'] if ct_name == '真人' else None
                    posts = dy[dy[cfg.content_tag_field] == tag] if tag else dy[dy[cfg.content_tag_field] != rules['小红书']['真人']['content_tag']]
                    if len(posts) > 0:
                        sub = self._calc_per_post(posts, rules['抖音'][ct_name],
                                                 rules['抖音']['max_posts'], cfg, label=f"抖音-{ct_name}")
                        if sub['小计'] > 0:
                            dy_total += sub['小计']
                            breakdown.append(sub)
                if dy_total > 0:
                    platforms['抖音'] = dy_total
                    total += dy_total

            # 快手
            ks = group[group[cfg.platform_field] == '快手']
            if len(ks) > 0:
                ks_total = 0.0
                for ct_name in ['真人', '非真人']:
                    tag = rules['小红书']['真人']['content_tag'] if ct_name == '真人' else None
                    posts = ks[ks[cfg.content_tag_field] == tag] if tag else ks[ks[cfg.content_tag_field] != rules['小红书']['真人']['content_tag']]
                    if len(posts) > 0:
                        sub = self._calc_per_post(posts, rules['快手'][ct_name],
                                                 rules['快手']['max_posts'], cfg, label=f"快手-{ct_name}")
                        if sub['小计'] > 0:
                            ks_total += sub['小计']
                            breakdown.append(sub)
                if ks_total > 0:
                    platforms['快手'] = ks_total
                    total += ks_total

            # B站/视频号
            bili = group[group[cfg.platform_field].isin(['B站', '微信视频号'])]
            if len(bili) > 0:
                sub = self._calc_bili(bili, rules['B站/视频号']['tiers'], cfg)
                if sub['小计'] > 0:
                    platforms['B站/视频号'] = sub['小计']
                    total += sub['小计']
                    breakdown.append(sub)

            if total > 0:
                capped = min(total, cfg.cap_per_person)
                creator_totals[cid] = {
                    '昵称': name, 'total': capped, 'before': total, **platforms,
                }
                creator_details[cid] = CreatorDetail(
                    creator_id=cid, name=name, total=capped, before_cap=total,
                    is_capped=total > cfg.cap_per_person,
                    platforms=platforms, breakdown=breakdown,
                )

        return creator_totals, creator_details

    def calculate(self, df: Optional[pd.DataFrame] = None) -> SettlementResult:
        """执行结算计算"""
        if df is None:
            df = self.result.raw_data
        if df is None:
            raise ValueError("请先调用 load_data() 加载底表")

        cfg = self.config
        rules = SETTLEMENT_RULES

        # 过滤非网易云 + 可结算
        valid_pool = df[
            (df[cfg.platform_field] != cfg.netease_platform) &
            (df['_can_settle'])
        ].copy()

        # 按网易云分发条件拆分为两组
        settle = valid_pool[valid_pool['_has_netease']].copy()
        no_netease = valid_pool[~valid_pool['_has_netease']].copy()

        # 结算有网易云的达人
        creator_totals, creator_details = self._settle_creators(settle, cfg, rules)

        # 计算无网易云达人的本应结算金额
        no_netease_totals, no_netease_details = self._settle_creators(no_netease, cfg, rules)

        # 统计（基于全部过审稿件池）
        guoshen_pool = valid_pool
        xhs_guoshen = len(guoshen_pool[guoshen_pool[cfg.platform_field] == '小红书'])
        boom_1000 = int((guoshen_pool[cfg.like_field] >= 1000).sum())
        total_non_netease = len(df[df[cfg.platform_field] != cfg.netease_platform])

        xhs_play = guoshen_pool[guoshen_pool[cfg.platform_field] == '小红书'][cfg.play_field].sum()
        other_play = guoshen_pool[~guoshen_pool[cfg.platform_field].isin(['小红书'])]['_play'].sum()
        est_exposure = (xhs_play if pd.notna(xhs_play) else 0) * 4 + (other_play if pd.notna(other_play) else 0)
        total_interact = int(guoshen_pool[cfg.interact_field].sum())

        grand_total = sum(c['total'] for c in creator_totals.values())
        grand_before = sum(c['before'] for c in creator_totals.values())

        stats = {
            'total_rows': self.result.stats.get('total_rows', len(df)),
            'netease_rows': self.result.stats.get('netease_rows', 0),
            'total_non_netease': total_non_netease,
            'guoshen_count': len(guoshen_pool),
            'xhs_guoshen': xhs_guoshen,
            'awarded_creators': len(creator_totals),
            'awarded_posts': sum(
                len([i for i in bd.get('items', []) if i.get('award', 0) > 0])
                for d in creator_details.values()
                for bd in d.breakdown
            ),
            'boom_1000': boom_1000,
            'grand_total': grand_total,
            'grand_before': grand_before,
            'capped_count': sum(1 for c in creator_totals.values() if c['before'] > cfg.cap_per_person),
            'estimated_exposure': est_exposure,
            'total_interact': total_interact,
            'guoshen_rate': len(guoshen_pool) / total_non_netease if total_non_netease > 0 else 0,
            'boom_rate': boom_1000 / xhs_guoshen if xhs_guoshen > 0 else 0,
        }

        self.result.creator_totals = creator_totals
        self.result.creator_details = creator_details
        self.result.no_netease_totals = no_netease_totals
        self.result.no_netease_details = no_netease_details
        self.result.no_netease_posts = no_netease
        self.result.stats.update(stats)
        self.result.settled_posts = settle

        # 未分发统计
        self.result.stats['no_netease_count'] = len(no_netease_totals)
        self.result.stats['no_netease_amount'] = sum(
            c['total'] for c in no_netease_totals.values()
        )

        return self.result

    def _calc_xhs(self, posts, cum_tiers, boom_tiers, boom_threshold, cfg, label=""):
        """小红书结算：累计点赞奖 + 单条爆款奖"""
        posts = posts.sort_values(cfg.like_field, ascending=False)
        boom = posts[posts[cfg.like_field] >= boom_threshold].head(3)
        non_boom = posts[~posts.index.isin(boom.index)]

        items = []
        boom_total = 0.0

        # 爆款奖
        for _, p in boom.iterrows():
            amt = self.apply_tier(p[cfg.like_field], boom_tiers)
            if amt > 0:
                items.append({
                    '作品ID': str(int(p[cfg.post_id_field])),
                    '点赞': int(p[cfg.like_field]),
                    'award': amt,
                    'type': '爆款奖',
                    'tier': self.get_tier_label(p[cfg.like_field], boom_tiers),
                })
                boom_total += amt

        # 累计奖
        cum_likes = int(non_boom.head(10)[cfg.like_field].sum())
        cum_amt = self.apply_tier(cum_likes, cum_tiers)
        if cum_amt > 0:
            cum_posts = []
            for _, p in non_boom.head(10).iterrows():
                if pd.notna(p[cfg.like_field]) and p[cfg.like_field] > 0:
                    cum_posts.append({
                        '作品ID': str(int(p[cfg.post_id_field])),
                        '点赞': int(p[cfg.like_field]),
                    })
            items.append({
                'type': '累计奖',
                'cum_likes': cum_likes,
                'cum_posts': cum_posts,
                'award': cum_amt,
                'tier': f"累计{cum_likes}赞 → {self.get_tier_label(cum_likes, cum_tiers)}",
            })

        return {
            'label': label,
            '小计': boom_total + cum_amt,
            '爆款奖金': boom_total,
            '累计奖金': cum_amt,
            '累计点赞': cum_likes,
            'items': items,
        }

    def _calc_per_post(self, posts, tiers, max_posts, cfg, label=""):
        """抖音/快手结算：逐条阶梯价"""
        p = posts.dropna(subset=['_play']).sort_values('_play', ascending=False).head(max_posts)
        items = []
        total = 0.0
        for _, r in p.iterrows():
            amt = self.apply_tier(r['_play'], tiers)
            items.append({
                '作品ID': str(int(r[cfg.post_id_field])),
                '播放量': int(r['_play']),
                'award': amt,
                'type': '阶梯奖',
                '达标': amt > 0,
                'tier': self.get_tier_label(r['_play'], tiers),
            })
            total += amt
        return {'label': label, '小计': total, 'items': items}

    def _calc_bili(self, posts, tiers, cfg):
        """B站/视频号结算：逐条点赞阈值"""
        p = posts.dropna(subset=[cfg.like_field])
        items = []
        total = 0.0
        for _, r in p.iterrows():
            amt = self.apply_tier(r[cfg.like_field], tiers)
            items.append({
                '作品ID': str(int(r[cfg.post_id_field])),
                '平台': r[cfg.platform_field],
                '点赞': int(r[cfg.like_field]),
                'award': amt,
                'type': '分发奖',
                'tier': self.get_tier_label(r[cfg.like_field], tiers),
            })
            total += amt
        return {'label': 'B站/视频号', '小计': total, 'items': items}

    def get_creator_detail(self, creator_id: str) -> Optional[CreatorDetail]:
        """查询单个达人结算明细"""
        return self.result.creator_details.get(creator_id)

    def compare_versions(self, old_path: str, new_path: str) -> dict:
        """对比新旧底表版本"""
        old_df = pd.read_excel(old_path, sheet_name='底表') if '底表' in \
            pd.ExcelFile(old_path).sheet_names else pd.read_excel(old_path)
        new_df = pd.read_excel(new_path, sheet_name='底表') if '底表' in \
            pd.ExcelFile(new_path).sheet_names else pd.read_excel(new_path)

        cfg = self.config
        old_ids = set(int(x) for x in old_df[cfg.post_id_field].dropna() if pd.notna(x))
        new_ids = set(int(x) for x in new_df[cfg.post_id_field].dropna() if pd.notna(x))

        added = new_ids - old_ids
        removed = old_ids - new_ids

        # 数据变更
        common = old_ids & new_ids
        changes = []
        for pid in list(common)[:500]:  # 限制检查数量
            try:
                orow = old_df[old_df[cfg.post_id_field] == pid]
                nrow = new_df[new_df[cfg.post_id_field] == pid]
                if len(orow) == 0 or len(nrow) == 0: continue
                orow, nrow = orow.iloc[0], nrow.iloc[0]
                changed = []
                for col in [cfg.play_field, cfg.like_field, cfg.content_tag_field,
                           cfg.review_field, cfg.access_field]:
                    ov = str(orow[col]) if pd.notna(orow[col]) else ''
                    nv = str(nrow[col]) if pd.notna(nrow[col]) else ''
                    if ov != nv:
                        changed.append(f"{col}: {ov}→{nv}")
                if changed:
                    changes.append({
                        '作品ID': str(pid),
                        '创作匠': str(nrow[cfg.creator_name_field]),
                        '创作匠ID': str(nrow[cfg.creator_id_field]),
                        '变更': '; '.join(changed),
                    })
            except:
                pass

        # 结算金额变化
        old_result = self.calculate(self.load_data(old_path))
        new_result = self.calculate(self.load_data(new_path))
        settlement_changes = []
        all_ids = set(list(old_result.creator_totals.keys()) + list(new_result.creator_totals.keys()))
        for cid in all_ids:
            old_amt = old_result.creator_totals.get(cid, {}).get('total', 0)
            new_amt = new_result.creator_totals.get(cid, {}).get('total', 0)
            if old_amt != new_amt:
                settlement_changes.append({
                    '创作匠ID': cid,
                    '旧版金额': old_amt,
                    '新版金额': new_amt,
                    '变化': new_amt - old_amt,
                })

        return {
            'added': len(added),
            'removed': len(removed),
            'modified': len(changes),
            'changes': changes,
            'settlement_changes': sorted(
                settlement_changes, key=lambda x: abs(x['变化']), reverse=True
            ),
        }

    def to_excel(self, output_path: str, include_details: bool = True):
        """导出结算结果为 Excel (v4)"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        import pandas as pd

        result = self.result
        cfg = self.config
        wb = Workbook()
        nf = Font(name='微软雅黑', size=10); bf = Font(name='微软雅黑', bold=True, size=10)
        sf = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        sfil = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ofil = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ca = Alignment(horizontal='center', vertical='center')
        la = Alignment(horizontal='left', vertical='center', wrap_text=True)
        mf = '#,##0'

        def wc(ws, r, c, v, font=nf, fmt=None, fill=None, align=ca):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = font; cell.border = tb; cell.alignment = align
            if fmt: cell.number_format = fmt
            if fill: cell.fill = fill

        def section(ws, r, cs, ce, text):
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
            for c in range(cs, ce+1): ws.cell(row=r, column=c).fill = sfil; ws.cell(row=r, column=c).border = tb
            ws.cell(row=r, column=cs).font = sf; ws.cell(row=r, column=cs).value = text

        s = result.stats; sorted_c = sorted(result.creator_totals.items(), key=lambda x: x[1]['total'], reverse=True)

        # Sheet 1: 结算明细
        ws1 = wb.active; ws1.title = '结算明细'
        ws1.merge_cells('A1:H1')
        ws1.cell(row=1, column=1, value=f'歌单推广任务结算 — {s["awarded_creators"]}人 ¥{s["grand_total"]:,.0f}').font = Font(name='微软雅黑', bold=True, size=14)
        row = 3; section(ws1, row, 1, 8, '创作者结算汇总'); row += 1
        for i, h in enumerate(['创作匠ID', '昵称', '小红书', '抖音', '快手', 'B站/视频号', '结算金额', '封顶前']):
            wc(ws1, row, 1+i, h, font=bf)
        row += 1
        for cid, data in sorted_c:
            wc(ws1, row, 1, cid); wc(ws1, row, 2, data['昵称'])
            for pi, p in enumerate(['小红书', '抖音', '快手', 'B站/视频号']):
                wc(ws1, row, 3+pi, data.get(p, 0), fmt=mf)
            wc(ws1, row, 7, data['total'], fmt=mf, font=bf)
            wc(ws1, row, 8, f"¥{data['before']:,.0f}" + (' 🔒封顶' if data['before'] > cfg.cap_per_person else '')); row += 1
        wc(ws1, row, 1, '', font=bf); wc(ws1, row, 2, '合计', font=bf)
        for pi, p in enumerate(['小红书', '抖音', '快手', 'B站/视频号']):
            wc(ws1, row, 3+pi, sum(d.get(p, 0) for _, d in sorted_c), fmt=mf, font=bf)
        wc(ws1, row, 7, s['grand_total'], fmt=mf, font=bf); row += 2
        section(ws1, row, 1, 8, '整体数据统计'); row += 1
        cpm = s['grand_total'] / s['estimated_exposure'] * 1000 if s.get('estimated_exposure', 0) > 0 else 0
        cpe = s['grand_total'] / s['total_interact'] if s.get('total_interact', 0) > 0 else 0
        for label, value in [('获奖人数', f'{s["awarded_creators"]}人'), ('封顶前/结算总金额', f'¥{s["grand_before"]:,.0f} / ¥{s["grand_total"]:,.0f}'), ('封顶人数', f'{s["capped_count"]}人'), ('过审条数', str(s.get("guoshen_count",""))), ('预估曝光cpm/CPE', f'{cpm:.2f} / {cpe:.2f}')]:
            wc(ws1, row, 1, '', font=bf); ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            wc(ws1, row, 2, label, font=bf, align=la); ws1.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            wc(ws1, row, 5, value); row += 1
        for ci in range(1, 9): ws1.column_dimensions[get_column_letter(ci)].width = 18

        # Sheet 2: 计算过程
        if include_details and result.creator_details:
            ws2 = wb.create_sheet('计算过程'); row = 1
            section(ws2, row, 1, 6, '达人结算计算明细'); row += 1
            for cid, detail in sorted(result.creator_details.items(), key=lambda x: x[1].total, reverse=True):
                wc(ws2, row, 1, cid, font=bf); wc(ws2, row, 2, detail.name, font=bf)
                wc(ws2, row, 3, f"¥{detail.total:,.0f}", font=bf); row += 1
                for bd in detail.breakdown:
                    wc(ws2, row, 1, bd['label'], font=Font(name='微软雅黑', bold=True, size=10, color='4472C4'))
                    wc(ws2, row, 2, f"¥{bd['小计']:,.0f}"); row += 1
                    for item in bd.get('items', []):
                        t = item.get('type', '')
                        if t == '爆款奖':
                            wc(ws2, row, 2, f"爆款 {item.get('作品ID','')[-8:]}")
                            wc(ws2, row, 3, f"{item.get('点赞','')}赞"); wc(ws2, row, 4, f"¥{item['award']:,.0f}")
                        elif t == '累计奖':
                            likes = '+'.join(str(p.get('点赞','')) for p in item.get('cum_posts', [])[:6])
                            wc(ws2, row, 2, '累计奖'); wc(ws2, row, 3, f"{item.get('cum_likes',0)}赞 ({likes})")
                            wc(ws2, row, 4, f"¥{item['award']:,.0f}")
                        elif t in ('阶梯奖', '分发奖'):
                            pl = item.get('播放量', item.get('点赞', ''))
                            wc(ws2, row, 2, f"{item.get('作品ID','')[-8:]}"); wc(ws2, row, 3, f"{pl:,}")
                            wc(ws2, row, 4, f"¥{item['award']:,.0f}")
                        wc(ws2, row, 5, item.get('tier','')); row += 1
                row += 1
            for ci in range(1, 7): ws2.column_dimensions[get_column_letter(ci)].width = 22

        # Sheet 3: 底表
        if result.raw_data is not None:
            ws3 = wb.create_sheet('底表')
            for ci, cn in enumerate(result.raw_data.columns): wc(ws3, 1, ci+1, str(cn), font=bf)
            for ri in range(min(len(result.raw_data), 2000)):
                for ci in range(len(result.raw_data.columns)):
                    val = result.raw_data.iloc[ri, ci]
                    if pd.isna(val): continue
                    try: wc(ws3, ri+2, ci+1, val)
                    except: pass

        # Sheet 4: 未分发网易云
        if result.no_netease_totals:
            ws4 = wb.create_sheet('未分发网易云达人')
            section(ws4, 1, 1, 6, '未分发网易云达人 — 本应结算金额'); row = 3
            for i, h in enumerate(['创作匠ID', '昵称', '投稿平台', '过审稿件数', '本应结算金额', '封顶前金额']):
                wc(ws4, row, 1+i, h, font=bf, fill=ofil)
            row += 1
            for cid, data in sorted(result.no_netease_totals.items(), key=lambda x: x[1]['total'], reverse=True):
                if data['total'] <= 0: continue
                wc(ws4, row, 1, cid); wc(ws4, row, 2, data['昵称'])
                wc(ws4, row, 3, ', '.join(k for k in data.keys() if k not in ('昵称', 'total', 'before')))
                nn = result.no_netease_posts[result.no_netease_posts[cfg.creator_id_field] == cid] if result.no_netease_posts is not None else pd.DataFrame()
                wc(ws4, row, 4, len(nn)); wc(ws4, row, 5, data['total'], fmt=mf, font=bf)
                wc(ws4, row, 6, data['before'], fmt=mf); row += 1
            for ci, w in enumerate([18, 18, 26, 12, 16, 16], 1): ws4.column_dimensions[get_column_letter(ci)].width = w

        wb.save(output_path)
        return output_path


